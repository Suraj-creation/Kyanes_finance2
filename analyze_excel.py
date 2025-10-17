import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import json

# File path
file_path = r"c:\Users\Govin\Desktop\finance\Copy of Final_Financial_Data_Kaynes_Technology(1).xlsx"

# Load the workbook
wb = openpyxl.load_workbook(file_path, data_only=False)

print("=" * 80)
print("EXCEL FILE COMPREHENSIVE ANALYSIS")
print("=" * 80)
print(f"\nFile: Copy of Final_Financial_Data_Kaynes_Technology(1).xlsx")
print(f"Total Sheets: {len(wb.sheetnames)}")
print(f"Sheet Names: {wb.sheetnames}")
print("\n")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    print("=" * 80)
    print(f"SHEET: {sheet_name}")
    print("=" * 80)
    
    ws = wb[sheet_name]
    
    # Basic sheet information
    print(f"\n📊 SHEET DIMENSIONS:")
    print(f"  - Max Row: {ws.max_row}")
    print(f"  - Max Column: {ws.max_column}")
    print(f"  - Used Range: A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    
    # Read data with pandas for better analysis
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        print(f"\n📋 DATA STRUCTURE:")
        print(f"  - Total Rows with Data: {len(df)}")
        print(f"  - Total Columns with Data: {len(df.columns)}")
        print(f"  - Non-empty Cells: {df.notna().sum().sum()}")
        
        # Display first few rows
        print(f"\n🔍 FIRST 10 ROWS OF DATA:")
        print(df.head(10).to_string())
        
        # Check for headers
        print(f"\n📑 POTENTIAL HEADERS (First Row):")
        print(df.iloc[0].to_dict() if len(df) > 0 else "No data")
        
        # Data types analysis
        print(f"\n🔢 DATA TYPES ANALYSIS:")
        for idx, col in enumerate(df.columns):
            non_null_count = df[col].notna().sum()
            unique_count = df[col].nunique()
            print(f"  Column {idx}: {non_null_count} non-null values, {unique_count} unique values")
        
        # Look for numeric data
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            print(f"\n💰 NUMERIC COLUMNS: {numeric_cols}")
            print(f"\nSTATISTICAL SUMMARY:")
            print(df[numeric_cols].describe().to_string())
        
        # Check for merged cells
        print(f"\n🔗 MERGED CELLS:")
        merged_cells = ws.merged_cells.ranges
        if merged_cells:
            for merged_cell in merged_cells:
                print(f"  - {merged_cell}")
        else:
            print("  - No merged cells found")
        
        # Check for formulas
        print(f"\n📐 FORMULAS DETECTED:")
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula_count += 1
                    if formula_count <= 10:  # Show first 10 formulas
                        print(f"  - {cell.coordinate}: {cell.value}")
        if formula_count > 10:
            print(f"  ... and {formula_count - 10} more formulas")
        if formula_count == 0:
            print("  - No formulas found")
        
        # Check for cell formatting
        print(f"\n🎨 CELL FORMATTING SAMPLE (First 5 rows):")
        for row_idx, row in enumerate(ws.iter_rows(max_row=5), 1):
            for cell in row:
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {cell.value} | Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
        
        # Check for comments/notes
        print(f"\n💬 COMMENTS/NOTES:")
        comment_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comment_count += 1
                    print(f"  - {cell.coordinate}: {cell.comment.text}")
        if comment_count == 0:
            print("  - No comments found")
        
        # Look for specific patterns (financial data)
        print(f"\n🏦 FINANCIAL DATA PATTERNS:")
        # Check for date patterns
        date_cols = df.select_dtypes(include=['datetime64']).columns.tolist()
        if date_cols:
            print(f"  - Date columns found: {date_cols}")
        
        # Check for specific keywords
        keywords = ['revenue', 'profit', 'loss', 'income', 'expense', 'assets', 'liability', 
                   'equity', 'cash', 'sales', 'cost', 'tax', 'depreciation', 'EBITDA', 'balance']
        
        found_keywords = []
        for keyword in keywords:
            if df.astype(str).apply(lambda x: x.str.contains(keyword, case=False, na=False)).any().any():
                found_keywords.append(keyword)
        
        if found_keywords:
            print(f"  - Financial keywords found: {', '.join(found_keywords)}")
        else:
            print("  - No common financial keywords detected in visible data")
        
        print("\n" + "-" * 80)
        
    except Exception as e:
        print(f"\n❌ Error reading sheet with pandas: {e}")
        print("Attempting to read raw cell values...")
        
        # Fallback: read raw values
        print(f"\n🔍 FIRST 10 ROWS (RAW):")
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            print(f"Row {row_idx}: {row}")
        
    print("\n")

# Additional workbook properties
print("=" * 80)
print("WORKBOOK PROPERTIES")
print("=" * 80)
print(f"\nWorkbook Properties:")
if wb.properties:
    print(f"  - Title: {wb.properties.title}")
    print(f"  - Creator: {wb.properties.creator}")
    print(f"  - Created: {wb.properties.created}")
    print(f"  - Modified: {wb.properties.modified}")
    print(f"  - Last Modified By: {wb.properties.lastModifiedBy}")

# Check for defined names (named ranges)
print(f"\n📛 DEFINED NAMES/RANGES:")
if wb.defined_names:
    for name in wb.defined_names.definedName:
        print(f"  - {name.name}: {name.value}")
else:
    print("  - No defined names found")

wb.close()
print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
