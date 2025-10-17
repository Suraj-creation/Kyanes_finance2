import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

file_path = r"c:\Users\Govin\Desktop\finance\Copy of Final_Financial_Data_Kaynes_Technology(1).xlsx"

# Read with pandas to see actual data
df = pd.read_excel(file_path, sheet_name='Table 3', header=None)

print("=" * 100)
print("KAYNES TECHNOLOGY FINANCIAL ANALYSIS - DETAILED COMPONENT BREAKDOWN")
print("=" * 100)

# Display all rows with actual content
print("\n📊 COMPLETE DATA STRUCTURE (ALL ROWS):\n")

for idx, row in df.iterrows():
    # Only show rows with some non-null values
    if row.notna().any():
        non_null_items = [(i, val) for i, val in enumerate(row) if pd.notna(val)]
        if non_null_items:
            print(f"\nRow {idx + 1}:")
            for col_idx, value in non_null_items:
                col_letter = get_column_letter(col_idx + 1)
                print(f"  [{col_letter}] {value}")

print("\n" + "=" * 100)
print("\n🔍 ANALYZING STRUCTURE - IDENTIFYING SECTIONS\n")

# Let's read it properly by identifying the sections
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Table 3']

sections = {
    'title': [],
    'cash_flow': [],
    'profit_loss': [],
    'balance_sheet': [],
    'ratios': []
}

print("Row-by-Row Content Analysis:\n")
for row_idx in range(1, min(ws.max_row + 1, 180)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            row_data.append(f"{get_column_letter(col_idx)}{row_idx}: {cell.value}")
    
    if row_data:
        print(f"\nRow {row_idx}: {' | '.join(row_data[:5])}")  # Show first 5 cells
        if len(row_data) > 5:
            print(f"         ... and {len(row_data) - 5} more cells")

wb.close()

print("\n" + "=" * 100)
print("DETAILED ANALYSIS COMPLETE")
print("=" * 100)
